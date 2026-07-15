import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


REGISTRY_ID = "coffee-origin-registry"
REGISTRY_VERSION = "2026.07.1"
PUBLISHED_AT = "2026-07-12T00:00:00Z"


def text(value):
    return "" if value is None else str(value).strip()


def rows_as_dicts(sheet):
    rows = sheet.iter_rows(values_only=True)
    headers = [text(value) for value in next(rows)]
    for values in rows:
        row = {headers[index]: values[index] for index in range(len(headers))}
        if any(value is not None and text(value) for value in values):
            yield row


def canonical_json(value):
    return json.dumps(
        value,
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")


def build_registry(workbook_path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        master_rows = sorted(
            rows_as_dicts(workbook["总编码表"]),
            key=lambda item: text(item["code"]),
        )
        qr_id_by_code = {
            text(item["code"]): index + 1
            for index, item in enumerate(master_rows)
        }

        entities = []
        for item in master_rows:
            code = text(item["code"])
            country_code = text(item["country_code"])
            entities.append({
                "qrId": qr_id_by_code[code],
                "code": code,
                "type": text(item["type"]),
                "countryQrId": qr_id_by_code.get(country_code),
                "nameCn": text(item["name_cn"]),
                "nameEn": text(item["name_en"]),
                "shortName": text(item["short_name"]),
                "group": text(item["group"]),
                "status": text(item["status"]) or "active",
                "introducedVersion": REGISTRY_VERSION,
                "sourceOrigin": text(item["source_origin"]),
                "sourceUrl": text(item["source_url"]),
            })

        relationships = []
        for row_number, item in enumerate(
            rows_as_dicts(workbook["关系映射表"]), start=2
        ):
            parent_code = text(item["parent_code"])
            child_code = text(item["child_code"])
            if parent_code not in qr_id_by_code or child_code not in qr_id_by_code:
                raise ValueError(
                    f"relationship row {row_number} references an unknown code"
                )
            relationships.append({
                "parentQrId": qr_id_by_code[parent_code],
                "childQrId": qr_id_by_code[child_code],
                "relationType": text(item["relation_type"]),
                "confidence": text(item["confidence"]),
                "sourceUrl": text(item["source_url"]),
            })
        relationships.sort(
            key=lambda item: (
                item["parentQrId"], item["childQrId"], item["relationType"]
            )
        )

        aliases = []
        for row_number, item in enumerate(
            rows_as_dicts(workbook["别名映射表"]), start=2
        ):
            code = text(item["code"])
            if code not in qr_id_by_code:
                raise ValueError(f"alias row {row_number} references an unknown code")
            aliases.append({
                "qrId": qr_id_by_code[code],
                "alias": text(item["alias_name"]),
                "language": text(item["language"]),
                "aliasType": text(item["alias_type"]),
                "uniqueTarget": text(item["is_unique_target"]).lower() == "yes",
            })
        aliases.sort(
            key=lambda item: (
                item["qrId"], item["language"], item["alias"], item["aliasType"]
            )
        )

        body = {
            "registryId": REGISTRY_ID,
            "registryVersion": REGISTRY_VERSION,
            "publishedAt": PUBLISHED_AT,
            "sourceWorkbook": workbook_path.name,
            "entities": entities,
            "relationships": relationships,
            "aliases": aliases,
        }
        registry_hash = f"sha256:{hashlib.sha256(canonical_json(body)).hexdigest()}"
        return {
            "registryId": REGISTRY_ID,
            "registryVersion": REGISTRY_VERSION,
            "registryHash": registry_hash,
            "publishedAt": PUBLISHED_AT,
            "sourceWorkbook": workbook_path.name,
            "entities": entities,
            "relationships": relationships,
            "aliases": aliases,
        }
    finally:
        workbook.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent.parent
        / "registry"
        / "coffee-origin-registry-2026.07.1.json",
    )
    args = parser.parse_args()

    workbook_path = args.workbook.resolve(strict=True)
    registry = build_registry(workbook_path)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(registry, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(json.dumps({
        "output": str(args.output.resolve()),
        "registryHash": registry["registryHash"],
        "entities": len(registry["entities"]),
        "relationships": len(registry["relationships"]),
        "aliases": len(registry["aliases"]),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
